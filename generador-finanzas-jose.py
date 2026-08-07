from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

VERDE = "1E5631"; VERDECLARO = "E3F0E6"; AMARILLO = "FFF2CC"; GRIS = "F5F5F5"; ROJO = "C0392B"; NARANJA = "FAD7A0"
BORDE = Side(style="thin", color="D9D9D9"); borde = Border(left=BORDE, right=BORDE, top=BORDE, bottom=BORDE)

wb = Workbook(); ws = wb.active; ws.title = "AGOSTO 2026"; ws.sheet_view.showGridLines = False

def H(c):
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    c.fill = PatternFill("solid", fgColor=VERDE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = borde

def C(c, money=False, center=False, amarillo=False, bold=False, rojo=False, small=False):
    c.font = Font(name="Arial", size=8 if small else 9, bold=bold, color=ROJO if rojo else "000000")
    c.border = borde
    if money: c.number_format = "$#,##0;($#,##0);-"
    if center: c.alignment = Alignment(horizontal="center")
    if amarillo: c.fill = PatternFill("solid", fgColor=AMARILLO)

def seccion(row, titulo):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    s = ws.cell(row=row, column=2, value=titulo)
    s.font = Font(name="Arial", bold=True, size=11, color=VERDE)
    s.fill = PatternFill("solid", fgColor=VERDECLARO)

def headers(row, cols):
    for i, h in enumerate(cols):
        x = ws.cell(row=row, column=2 + i, value=h); H(x)

ws.merge_cells("B2:H2"); t = ws["B2"]; t.value = "FINANZAS DE JOSÉ — AGOSTO 2026"
t.font = Font(name="Arial", bold=True, size=15, color=VERDE)
ws.merge_cells("B3:H3"); s = ws["B3"]
s.value = "Sueldo $1.800.000 · TC editable (naranja) · Jarvis actualiza con cada gasto/comprobante · UNA sola planilla, siempre"
s.font = Font(name="Arial", size=8, italic=True, color="808080")
ws["G4"].value = "TC $/USD:"; ws["G4"].font = Font(name="Arial", bold=True, size=9)
tc = ws["H4"]; tc.value = 1520; tc.number_format = "$#,##0.00"; tc.fill = PatternFill("solid", fgColor=NARANJA); tc.border = borde; tc.font = Font(name="Arial", bold=True, size=9)

r = 6
seccion(r, "🔴 PAGAR YA — VISA GOLD BBVA (vence JUE 7/8)")
headers(r + 1, ["Concepto", "", "Monto", "", "", "", "Notas"])
pagos = [
    ("Pago en PESOS (saldo − percep. RG5617 $56.356)", 2051525, "NUNCA financiar · pagar total como siempre"),
    ("Comprar USD 125,57 y pagar en dólares", "=ROUND(125.57*$H$4,0)", "Recupera la percep. el mes próximo · Apple ya cobrado a Facu"),
    ("Entrenador personal (efectivo, al cobrar)", 100000, "Todos los meses"),
]
f0 = r + 2
for i, (n, m, nota) in enumerate(pagos):
    rr = f0 + i
    C(ws.cell(row=rr, column=2, value=n), bold=True)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
    C(ws.cell(row=rr, column=4, value=m), money=True, bold=True)
    ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=8)
    C(ws.cell(row=rr, column=7, value=nota), small=True)

r = f0 + len(pagos) + 1
seccion(r, "🟢 PLATA QUE ENTRA / YA COBRADA")
headers(r + 1, ["Concepto", "", "Monto", "Estado", "", "", "Notas"])
ingresos = [
    ("Sueldo", 1800000, "por cobrar", ""),
    ("Facu — Mutual Sancor", 243726, "COBRADO ✓", ""),
    ("Facu — Smash Pádel", 212667, "COBRADO ✓", ""),
    ("Facu — Apple USD 15,97 (TC 1520)", 24274, "COBRADO ✓", ""),
    ("Agustín — plan de pago (débito día 16)", 58000, "cobrar c/mes", "Tener plata en la cuenta el 16 para el débito"),
    ("Agustín — cuota moto ASPEN", 212000, "cobrar", "⚠️ La cuota NO vino en este resumen — investigar"),
    ("ZIBA (estética) — USD 900", "=ROUND(900*$H$4,0)", "POTENCIAL", "Si sale rápido salva el mes"),
    ("VENTA MOTO (piso autorizado)", 10200000, "POTENCIAL", "Restar gestor $40.000 + plan patentes DGR"),
]
f0 = r + 2
for i, (n, m, e, nota) in enumerate(ingresos):
    rr = f0 + i
    C(ws.cell(row=rr, column=2, value=n), bold=True)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
    C(ws.cell(row=rr, column=4, value=m), money=True)
    C(ws.cell(row=rr, column=5, value=e), center=True, small=True)
    ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=8)
    C(ws.cell(row=rr, column=7, value=nota), small=True)

r = f0 + len(ingresos) + 1
seccion(r, "📌 DEUDAS PERSONALES")
headers(r + 1, ["A quién", "", "Monto", "Cuándo", "", "", "Notas"])
deudas = [
    ("Dentista — corona (saldo)", 350000, "EXIGE YA", "Factura $700k presentada a OSDE (reintegro 50%) — SEGUIR reintegro; si OSDE paga, cubre el saldo"),
    ("Dentista — alineadores (6 cambios)", 240000, "en curso", "3ª etapa: ~15 alineadores, $40.000 cada 15-20 días (recurrente)"),
    ("Dentista — saldo inicial tratamiento", "=ROUND(1000*$H$4,0)", "al finalizar", "USD 1.000 · ir cancelando por partes"),
]
f0 = r + 2
for i, (n, m, e, nota) in enumerate(deudas):
    rr = f0 + i
    C(ws.cell(row=rr, column=2, value=n), bold=True)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
    C(ws.cell(row=rr, column=4, value=m), money=True)
    C(ws.cell(row=rr, column=5, value=e), center=True, small=True)
    ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=8)
    C(ws.cell(row=rr, column=7, value=nota), small=True)

r = f0 + len(deudas) + 1
seccion(r, "🔁 GASTOS FIJOS DE TODOS LOS MESES (cashflow)")
headers(r + 1, ["Concepto", "", "Monto/mes", "", "", "", "Notas"])
fijos = [
    ("Nafta camioneta ($50.000 x miércoles)", 200000, "también en resumen como Shellbox"),
    ("Nafta moto (hasta venderla)", 20000, "desaparece con la venta"),
    ("OSDE propia", 294077, "pagada con tarjeta = se financia 1 mes"),
    ("Entrenador personal (efectivo)", 100000, ""),
    ("Suscripciones USD (~110: Google, Claude…)", "=ROUND(110*$H$4,0)", "Apple es de Facu — cobrarle cada mes"),
    ("Alineadores (~2 cambios/mes)", 80000, "$40.000 cada 15-20 días"),
]
f0 = r + 2
for i, (n, m, nota) in enumerate(fijos):
    rr = f0 + i
    C(ws.cell(row=rr, column=2, value=n), bold=True)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
    C(ws.cell(row=rr, column=4, value=m), money=True)
    ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=8)
    C(ws.cell(row=rr, column=7, value=nota), small=True)

r = f0 + len(fijos) + 1
seccion(r, "📅 CUOTAS QUE VIENEN (tarjeta + compromisos)")
headers(r + 1, ["Mes", "", "Tarjeta (ya cerrado)", "Obra FAGU (nuevo)", "Aspen Agus (si vuelve)", "", "Notas"])
cuotas = [
    ("SEPTIEMBRE", 274343, 140454, 212000, "TMC $78.651 + Marola $61.803 · ⚠️ verificar ANULACIÓN Marola 1 cuota"),
    ("OCTUBRE", 274343, 140454, 212000, ""),
    ("NOVIEMBRE", 45000, 140454, 212000, ""),
    ("DICIEMBRE", 45000, 0, 212000, ""),
    ("ENERO 27", 45000, 0, 212000, ""),
]
f0 = r + 2
for i, (n, t1, t2, t3, nota) in enumerate(cuotas):
    rr = f0 + i
    C(ws.cell(row=rr, column=2, value=n), bold=True)
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
    C(ws.cell(row=rr, column=4, value=t1), money=True)
    C(ws.cell(row=rr, column=5, value=t2), money=True)
    C(ws.cell(row=rr, column=6, value=t3), money=True)
    ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=8)
    C(ws.cell(row=rr, column=7, value=nota), small=True)

anchos = {"A": 2, "B": 30, "C": 10, "D": 15, "E": 15, "F": 15, "G": 22, "H": 22}
for col, w in anchos.items(): ws.column_dimensions[col].width = w

g = wb.create_sheet("Guía"); g.sheet_view.showGridLines = False
g.column_dimensions["B"].width = 28; g.column_dimensions["C"].width = 66
tt = g["B2"]; tt.value = "CÓMO FUNCIONA"; tt.font = Font(name="Arial", bold=True, size=14, color=VERDE)
guia = [
    ("Regla de la tarjeta", "SIEMPRE pago total (nunca financiar al 94% TEA). Pesos con percepción RG5617 descontada + la parte USD comprada y pagada en dólares → devuelven la percep. el mes siguiente."),
    ("Préstamos a hermanos", "Todo gasto de Facu/Agus en las tarjetas de José se anota acá y se cobra: Sancor, Smash, Apple (Facu ✓) · Aspen 12 cuotas y plan del día 16 (Agus)."),
    ("Día 16", "Débito automático ~$58.000 (plan de Agustín): tener plata en la caja de ahorro y cobrárselo a Agus."),
    ("Al vender la moto", "Sumar: gestor ~$40.000 + plan de pago contado DGR (patentes). Desaparece nafta moto $20.000/mes. Entra ~$10,2M."),
    ("Los gastos del día a día", "Se cargan en la APP de gastos (José/Facu/Agus) — proyecto Personal / Tafí / Obra FAGU. Jarvis los baja acá y al cashflow."),
    ("Una sola planilla", "Esta es LA planilla de José. Jarvis edita in-place. Cada mes se agrega una hoja nueva."),
]
rg = 4
for k, v in guia:
    g.cell(row=rg, column=2, value=k).font = Font(name="Arial", bold=True, size=10)
    cv = g.cell(row=rg, column=3, value=v); cv.font = Font(name="Arial", size=10)
    cv.alignment = Alignment(wrap_text=True, vertical="top"); rg += 1

out = "/private/tmp/claude-501/-Users-josecrespin-Downloads/6673703f-6225-4fa9-ab05-2b0a08b147a1/scratchpad/FINANZAS-JOSE-2026.xlsx"
wb.save(out); print("OK")
